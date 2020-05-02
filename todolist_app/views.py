from django.shortcuts import render, redirect
from django.http import HttpResponse
from todolist_app.models import TaskList
from todolist_app.models import Gstworker
from todolist_app.forms import TaskForm
from django.contrib import messages
from django.core.paginator import Paginator
from todolist_app.functions.functions import handle_uploaded_file 
from todolist_app.forms import StudentForm
import csv
import json
#import xlsxwriter
from datetime import datetime
from datetime import timedelta
#from xlsxwriter import Workbook
from openpyxl import Workbook
from zipfile import ZipFile
import zipfile





def index1aaa(request):  # Main Code
     if request.method == 'POST':
          r_count = 0
          i = 0 # stands for number of GSTIN in B2B records
          j = 2 # Stands for Row 2 indicates data write is gonna start from 2nd row 
          k = 0 # Stands for count of invoices in a GSTIN Record
          l = 0 # Stands for number of invoice line items in a invoice record
          form = StudentForm(request.POST, request.FILES)  
          client_file = request.FILES['file']
          files = request.FILES.getlist('file')
          response = HttpResponse(
          content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
          )
          response['Content-Disposition'] = 'attachment; filename={date}-CA Ram Report Ver1_0_1.xlsx'.format(
          date=datetime.now().strftime('%Y-%m-%d'),
          )
          workbook = Workbook()
          worksheet = workbook.active
          worksheet.title = 'B2B'
 
      # unzip the zip file to the same directory 
          if form.is_valid(): 
               for f in files:
                    handle_uploaded_file(f)
                    with zipfile.ZipFile(f, 'r') as zip_ref:
                         first = zip_ref.infolist()[0]
                         with zip_ref.open(first, "r") as fo:
                              a = json.load(fo)
                    if form.is_valid():
                         try:
                              l = 0
                              k = 0
                              i = 0
                              j += 1 
                              while i < (len(a['b2b'])):
                                   while k < len(a['b2b'][i]['inv']):
                                        while l < (len(a['b2b'][i]['inv'][k]['itms'])):
                                             #worksheet.cell(row = j, column = 10).value = (a['b2b'][i]['inv'][k]['itms'][l]['num'])
                                             try:
                                                  worksheet.cell(row = j, column = 8).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['rt'])
                                             except:
                                                  pass
                                             try:
                                                  worksheet.cell(row = j, column = 9).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['txval'])
                                             except:
                                                  pass
                                             try:
                                                  worksheet.cell(row = j, column = 10).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['iamt'])
                                             except:
                                                  pass
                                             try:
                                                  worksheet.cell(row = j, column = 11).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['camt'])
                                             except:
                                                  pass        
                                             try: 
                                                  worksheet.cell(row = j, column = 12).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['samt'])
                                             except:
                                                  pass 
                                             try: 
                                                  worksheet.cell(row = j, column = 13).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['csamt'])
                                             except:
                                                  pass   
                                             try:
                                                  worksheet.cell(row = j, column = 1).value = (a['b2b'][i]['ctin'])
                                             except:
                                                  pass
                                             try:
                                                  worksheet.cell(row = j, column = 2).value = (a['b2b'][i]['inv'][k]['val'])
                                             except:
                                                  pass
                                             try:
                                                  worksheet.cell(row = j, column = 3).value = (a['b2b'][i]['inv'][k]['inv_typ'])
                                             except:   
                                                  pass
                                             try: 
                                                  worksheet.cell(row = j, column = 4).value = (a['b2b'][i]['inv'][k]['pos'])
                                             except:
                                                  pass
                                             try:
                                                  worksheet.cell(row = j, column = 5).value = (a['b2b'][i]['inv'][k]['idt'])
                                             except:
                                                  pass
                                             try:
                                                  worksheet.cell(row = j, column = 6).value = (a['b2b'][i]['inv'][k]['rchrg'])
                                             except:
                                                  pass
                                             try:
                                                  worksheet.cell(row = j, column = 7).value = (a['b2b'][i]['inv'][k]['inum'])
                                             except:
                                                  pass
                                             try:
                                                  worksheet.cell(row = j, column = 14).value = (a['gstin'])
                                             except:
                                                  pass
                                             try:
                                                  worksheet.cell(row = j, column = 15).value = (a['fp'])
                                             except:
                                                  pass
                                             
                                             r_count += 1
                                             l += 1 # Refers to callout the next invoice level line item hope it starts with 0
                                             j += 1 # Excel offset move to next row    
                                        l = 0 # Resetting to 0 for a new record 
                                        k += 1 # Refers to callout next invoice item for a gst record
                                   i += 1 # Moving to next GSTIN
                                   k = 0 # Resetting to 0 for a new record of Invoice
                         except:
                              pass  
                         #handle_uploaded_file(request.FILES['file'])
                         #return HttpResponse("File uploaded successfuly" ) 
                    workbook.save(response)
          return response     
     else:  
          student = StudentForm()  
          return render(request,"index.html",{'form':student})       
          return self.form_valid(form)     
# Create your views here.
def todolist(request):
          if request.method == "POST":
               form = TaskForm(request.POST or None)
               if form.is_valid():
                    form.save()
               messages.success(request,("New Task Added Successfully!"))     
               return  redirect('todolist')
          else:
               all_tasks = TaskList.objects.all()
               paginator = Paginator(all_tasks, 10)
               page = request.GET.get('pg')
               all_tasks = paginator.get_page(page)
               
               return render(request, 'todolist.html', {'all_tasks': all_tasks}) 
def delete_task(request, task_id):
          task = TaskList.objects.get(pk=task_id)
          task.delete()
          return  redirect('todolist')
def edit_task(request, task_id):
     if request.method == "POST":
          if request.method == "POST":
               task = TaskList.objects.get(pk=task_id)
               form = TaskForm(request.POST or None, instance=task)
               if form.is_valid():
                    form.save()
          messages.success(request,("Task Edited "))     
          return  redirect('todolist')
     else:
          task_obj = TaskList.objects.get(pk=task_id)
          return render(request, 'edit.html', {'task_obj': task_obj})
def contact(request):
     context = {
          'Contact_text' : " Welcome to  Contact Page.",
          }
     return render(request, 'contact.html', context)
def complete_task(request, task_id):
          task = TaskList.objects.get(pk=task_id)
          task.done = True
          task.save()
          return  redirect('todolist')
def pending_task(request, task_id):
          task = TaskList.objects.get(pk=task_id)
          task.done = False
          task.save()
          return  redirect('todolist')
          
def about(request):
     context = {
          'Welcome_text' : " Welcome About Page.",
          }
     return render(request, 'about.html', context)
def index(request):
     context = {
          'index_text' : " Welcome to  Index Page.",
          }
     return render(request, 'index.html', context)   
def index2(request):  # testing to extract multiple zip files
     if request.method == 'POST':  
          #student = StudentForm(request.POST, request.FILES)
          
          client_file = request.FILES['file']
                # unzip the zip file to the same directory 
          with zipfile.ZipFile(client_file, 'r') as zip_ref:
                    first = zip_ref.infolist()[0]
                    with zip_ref.open(first, "r") as fo:
                        a = json.load(fo)
          return HttpResponse( len(a['b2b']))
     else:  
          student = StudentForm()  
          return render(request,"index2.html",{'form':student})


    # Create your views here.

from django.views.generic.edit import FormView
from .forms import FileFieldForm

class FileFieldView(FormView):
    form_class = FileFieldForm
    template_name = 'index.html'  # Replace with your template.
    #success_url = 'todolist.html' # Replace with your URL or reverse().

    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        files = request.FILES.getlist('file_field')
        if form.is_valid():
            for f in files:
                ...  # Do something with each file.
            return self.form_valid(form)
        else:
            return self.form_invalid(form)




def index1a(request):  # Main Code
     if request.method == 'POST':
          form = StudentForm(request.POST, request.FILES)  
          client_file = request.FILES['file']
          files = request.FILES.getlist('file')
          #handle_uploaded_file(request.FILES['file'])
#return HttpResponse("File uploaded successfuly" ) 
          response = HttpResponse(
               content_type='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8',
          )
          response['Content-Disposition'] = 'attachment; filename={date}-CA Ram Report Ver1_0_1.xlsx'.format(
               date=datetime.now().strftime('%Y-%m-%d'),
          )
          workbook = Workbook()
# Get active worksheet/tab

          ws_Info = workbook.active
          ws_Info.title = 'Info'
          worksheet = workbook.create_sheet("B2B") # insert at the end (default)
          worksheet.cell(row = 1, column = 1).value = "Customer GSTIN"
          worksheet.cell(row = 1, column = 2).value = "Total Invoice Value"
          worksheet.cell(row = 1, column = 3).value = "Type of Invoice"
          worksheet.cell(row = 1, column = 4).value = "Place of Supply"
          worksheet.cell(row = 1, column = 5).value = "Date of Invoice"
          worksheet.cell(row = 1, column = 6).value = "Rcm Applicable"
          worksheet.cell(row = 1, column = 7).value = "Invoice Number"
          worksheet.cell(row = 1, column = 8).value = "Rate"
          worksheet.cell(row = 1, column = 9).value = "Taxable Value"
          worksheet.cell(row = 1, column = 10).value = "IGST"
          worksheet.cell(row = 1, column = 11).value = "CGST"
          worksheet.cell(row = 1, column = 12).value = "SGST"
          worksheet.cell(row = 1, column = 13).value = "CESS"
          worksheet.cell(row = 1, column = 14).value = "Dealer GSTIN"
          worksheet.cell(row = 1, column = 15).value = "Filing Period"
          ws_B2CL = workbook.create_sheet("B2CL")
          ws_B2CL.cell(row = 1, column = 1).value  = "Invoice Number"
          ws_B2CL.cell(row = 1, column = 2).value  = "Date of Invoice"
          ws_B2CL.cell(row = 1, column = 3).value  = "Total Invoice Value"
          ws_B2CL.cell(row = 1, column = 4).value  = "Place of Supply"
          ws_B2CL.cell(row = 1, column = 5).value  = "Rate"
          ws_B2CL.cell(row = 1, column = 6).value  = "Taxable Value"
          ws_B2CL.cell(row = 1, column = 7).value  = "IGST"
          ws_B2CL.cell(row = 1, column = 8).value  = "CGST"
          ws_B2CL.cell(row = 1, column = 9).value  = "SGST"
          ws_B2CL.cell(row = 1, column = 10).value  = "CESS"
          ws_B2CL.cell(row = 1, column = 11).value  = "Diff % Tax Rate (If Any)"
          ws_B2CL.cell(row = 1, column = 12).value  = "Section 7 Supplies ?"
          ws_B2CL.cell(row = 1, column = 13).value  = "Dealer GSTIN"
          ws_B2CL.cell(row = 1, column = 14).value  = "Filing Period"
          ws_B2BA = workbook.create_sheet("B2BA")
          ws_B2BA.cell(row = 1, column = 1).value  = "Customer GSTIN"
          ws_B2BA.cell(row = 1, column = 2).value  = "Old Invoice Number" 
          ws_B2BA.cell(row = 1, column = 3).value  = "Old Invoice Date" 
          ws_B2BA.cell(row = 1, column = 4).value  = "Invoice Number" 
          ws_B2BA.cell(row = 1, column = 5).value  = "Invoice Date" 
          ws_B2BA.cell(row = 1, column = 6).value  = "Total Invoice Value" 
          ws_B2BA.cell(row = 1, column = 7).value  = "Place of Supply" 
          ws_B2BA.cell(row = 1, column = 8).value  = "RCM Applicability" 
          ws_B2BA.cell(row = 1, column = 9).value  = "Diff % Tax Rate" 
          ws_B2BA.cell(row = 1, column = 10).value  = "Invoice Type" 
          ws_B2BA.cell(row = 1, column = 11).value  = "Taxable Value" 
          ws_B2BA.cell(row = 1, column = 12).value  = "Rate" 
          ws_B2BA.cell(row = 1, column = 13).value  = "IGST" 
          ws_B2BA.cell(row = 1, column = 14).value  = "CGST" 
          ws_B2BA.cell(row = 1, column = 15).value  = "SGST" 
          ws_B2BA.cell(row = 1, column = 16).value  = "CESS"
          ws_B2BA.cell(row = 1, column = 17).value  = "Dealer GSTIN"
          ws_B2BA.cell(row = 1, column = 18).value  = "Filing Period"         


          ws_B2CLA = workbook.create_sheet("B2CLA")
          ws_B2CLA.cell(row = 1, column = 1).value  = "Place of Supply"
          ws_B2CLA.cell(row = 1, column = 2).value  = "Old Invoice Number"
          ws_B2CLA.cell(row = 1, column = 3).value  = "Old invoice Date"
          ws_B2CLA.cell(row = 1, column = 4).value  = "Revised Invoice Number"
          ws_B2CLA.cell(row = 1, column = 5).value  = "Revised Invoice Date"
          ws_B2CLA.cell(row = 1, column = 6).value  = "Total Invoice Value"
          ws_B2CLA.cell(row = 1, column = 7).value  = "Diff % Tax Rate"
          ws_B2CLA.cell(row = 1, column = 8).value  = "Invoice Type"
          ws_B2CLA.cell(row = 1, column = 9).value  = "Taxable Value"
          ws_B2CLA.cell(row = 1, column = 10).value  = "Rate"
          ws_B2CLA.cell(row = 1, column = 11).value  = "IGST"
          ws_B2CLA.cell(row = 1, column = 12).value  = "CGST"
          ws_B2CLA.cell(row = 1, column = 13).value  = "SGST"
          ws_B2CLA.cell(row = 1, column = 14).value  = "Cess"
          ws_B2CLA.cell(row = 1, column = 14).value  = "Dealer GSTIN"
          ws_B2CLA.cell(row = 1, column = 14).value  = "Filing Period"

          ws_B2CS = workbook.create_sheet("B2CS")
          ws_B2CS.cell(row = 1, column = 1).value  = "Supply Type"
          ws_B2CS.cell(row = 1, column = 2).value  = "Rate"
          ws_B2CS.cell(row = 1, column = 3).value  = "E-Commerce Supply"
          ws_B2CS.cell(row = 1, column = 4).value  = "Place of Supply"
          ws_B2CS.cell(row = 1, column = 5).value  = "Differential Tax Rate"   
          ws_B2CS.cell(row = 1, column = 6).value  = "Taxable Value"
          ws_B2CS.cell(row = 1, column = 7).value  = "IGST"
          ws_B2CS.cell(row = 1, column = 8).value  = "CGST"
          ws_B2CS.cell(row = 1, column = 9).value  = "SGST"
          ws_B2CS.cell(row = 1, column = 10).value  = "CESS"
          ws_B2CS.cell(row = 1, column = 11).value  = "Dealer GSTIN"
          ws_B2CS.cell(row = 1, column = 12).value  = "Filing Period"

          ws_B2CSA = workbook.create_sheet("B2CSA")
          ws_B2CSA.cell(row = 1, column = 1).value  = "Original Month"
          ws_B2CSA.cell(row = 1, column = 2).value  = "Supply Type"
          ws_B2CSA.cell(row = 1, column = 3).value  = "E-Comerce Supply?"
          ws_B2CSA.cell(row = 1, column = 4).value  = "Place of Supply"
          ws_B2CSA.cell(row = 1, column = 5).value  = "Differential Tax Rate"   
          ws_B2CSA.cell(row = 1, column = 6).value  = "Taxable Value"
          ws_B2CSA.cell(row = 1, column = 7).value  = "IGST"
          ws_B2CSA.cell(row = 1, column = 8).value  = "CGST"
          ws_B2CSA.cell(row = 1, column = 9).value  = "SGST"
          ws_B2CSA.cell(row = 1, column = 10).value  = "CESS"
          ws_B2CSA.cell(row = 1, column = 11).value  = "Dealer GSTIN"
          ws_B2CSA.cell(row = 1, column = 12).value  = "Filing Period"


          ws_EXP = workbook.create_sheet("EXP")
          ws_EXP.cell(row = 1, column = 1).value  = "Export Type"
          ws_EXP.cell(row = 1, column = 2).value  = "Invoice Number"
          ws_EXP.cell(row = 1, column = 3).value  = "Invoice Date"
          ws_EXP.cell(row = 1, column = 4).value  = "Invoice Value"
          ws_EXP.cell(row = 1, column = 5).value  = "Port Code"
          ws_EXP.cell(row = 1, column = 6).value  = "Shipping Bill Number"
          ws_EXP.cell(row = 1, column = 7).value  = "Shipping Bill Date"
          ws_EXP.cell(row = 1, column = 8).value  = "Taxable Value"
          ws_EXP.cell(row = 1, column = 9).value  = "Rate"
          ws_EXP.cell(row = 1, column = 10).value  = "IGST"
          ws_EXP.cell(row = 1, column = 11).value  = "CESS"
          ws_EXP.cell(row = 1, column = 12).value  = "Dealer GSTIN"
          ws_EXP.cell(row = 1, column = 13).value  = "Filing Period"

          ws_EXPA = workbook.create_sheet("EXPA")
          ws_EXPA.cell(row = 1, column = 1).value  = "Export Type"
          ws_EXPA.cell(row = 1, column = 2).value  = "Invoice Number"
          ws_EXPA.cell(row = 1, column = 3).value  = "Invoice Date"
          ws_EXPA.cell(row = 1, column = 4).value  = "Invoice Value"
          ws_EXPA.cell(row = 1, column = 5).value  = "Port Code"
          ws_EXPA.cell(row = 1, column = 6).value  = "Shipping Bill Number"
          ws_EXPA.cell(row = 1, column = 7).value  = "Shipping Bill Date"
          ws_EXPA.cell(row = 1, column = 8).value  = "Taxable Value"
          ws_EXPA.cell(row = 1, column = 9).value  = "Rate"
          ws_EXPA.cell(row = 1, column = 10).value  = "IGST"
          ws_EXPA.cell(row = 1, column = 11).value  = "CESS"  
          ws_EXPA.cell(row = 1, column = 12).value  = "Old Invoice Number"  
          ws_EXPA.cell(row = 1, column = 13).value  = "Old Invoice Date"  
          ws_EXPA.cell(row = 1, column = 14).value  = "Dealer GSTIN"
          ws_EXPA.cell(row = 1, column = 15).value  = "Filing Period"
          ws_CDNR = workbook.create_sheet("CDNR")
          ws_CDNR.cell(row = 1, column = 1).value  = "Customer GSTIN"
          ws_CDNR.cell(row = 1, column = 2).value  = "Note Value"
          ws_CDNR.cell(row = 1, column = 3).value  = "Note Type"
          ws_CDNR.cell(row = 1, column = 4).value  = "Note Number"
          ws_CDNR.cell(row = 1, column = 5).value  = "Invoice Number"
          ws_CDNR.cell(row = 1, column = 6).value  = "Invoice Date"
          ws_CDNR.cell(row = 1, column = 7).value  = "Note Date"
          ws_CDNR.cell(row = 1, column = 8).value  = "Is Pre GST ?"
          ws_CDNR.cell(row = 1, column = 9).value  = "Taxable Value"
          ws_CDNR.cell(row = 1, column = 10).value  = "Rate"
          ws_CDNR.cell(row = 1, column = 11).value  = "IGST"
          ws_CDNR.cell(row = 1, column = 12).value  = "CGST"
          ws_CDNR.cell(row = 1, column = 13).value  = "SGST"
          ws_CDNR.cell(row = 1, column = 14).value  = "CESS"
          ws_CDNR.cell(row = 1, column = 15).value  = "Dealer GSTIN"
          ws_CDNR.cell(row = 1, column = 16).value  = "Filing Period"
          
          ws_CDNRA = workbook.create_sheet("CDNRA")
          ws_CDNRA.cell(row = 1, column = 1).value  = "Customer GSTIN"
          ws_CDNRA.cell(row = 1, column = 2).value  = "Note Value"
          ws_CDNRA.cell(row = 1, column = 3).value  = "Note Type"
          ws_CDNRA.cell(row = 1, column = 4).value  = "Old Note Number"
          ws_CDNRA.cell(row = 1, column = 5).value  = "Old Note Date"
          
          ws_CDNRA.cell(row = 1, column = 6).value  = "Note Number"
          ws_CDNRA.cell(row = 1, column = 7).value  = "Invoice Number"
          ws_CDNRA.cell(row = 1, column = 8).value  = "Invoice Date"
          ws_CDNRA.cell(row = 1, column = 9).value  = "Note Date"
          ws_CDNRA.cell(row = 1, column = 10).value  = "Is Pre GST ?"
          ws_CDNRA.cell(row = 1, column = 11).value  = "Taxable Value"
          ws_CDNRA.cell(row = 1, column = 12).value  = "Rate"
          ws_CDNRA.cell(row = 1, column = 13).value  = "IGST"
          ws_CDNRA.cell(row = 1, column = 14).value  = "CGST"
          ws_CDNRA.cell(row = 1, column = 15).value  = "SGST"
          ws_CDNRA.cell(row = 1, column = 16).value  = "CESS"
          ws_CDNRA.cell(row = 1, column = 17).value  = "Dealer GSTIN"
          ws_CDNRA.cell(row = 1, column = 18).value  = "Filing Period"


          ws_CDNUR = workbook.create_sheet("CDNUR")

          ws_CDNUR.cell(row = 1, column = 1).value  = "UR TYPE"
          ws_CDNUR.cell(row = 1, column = 2).value  = "Note Number"
          ws_CDNUR.cell(row = 1, column = 3).value  = "Note Date"
          ws_CDNUR.cell(row = 1, column = 4).value  = "Invoice Number"
          ws_CDNUR.cell(row = 1, column = 5).value  = "Invoice Date"
          ws_CDNUR.cell(row = 1, column = 6).value  = "Note Type"
          ws_CDNUR.cell(row = 1, column = 7).value  = "Is Pre GST?"
          ws_CDNUR.cell(row = 1, column = 8).value  = "Note Value"
          ws_CDNUR.cell(row = 1, column = 9).value  = "Taxable Value"
          ws_CDNUR.cell(row = 1, column = 10).value  = "Rate"
          ws_CDNUR.cell(row = 1, column = 11).value  = "IGST"
          ws_CDNUR.cell(row = 1, column = 12).value  = "CGST"
          ws_CDNUR.cell(row = 1, column = 13).value  = "SGST"
          ws_CDNUR.cell(row = 1, column = 14).value  = "CESS"
          ws_CDNUR.cell(row = 1, column = 15).value  = "Dealer GSTIN"
          ws_CDNUR.cell(row = 1, column = 16).value  = "Filing Period"
          ws_CDNURA = workbook.create_sheet("CDNURA")
          ws_CDNURA.cell(row = 1, column = 1).value  = "UR TYPE"
          ws_CDNURA.cell(row = 1, column = 2).value  = "Old Note Number"
          ws_CDNURA.cell(row = 1, column = 3).value  = "Old Note Date"
          ws_CDNURA.cell(row = 1, column = 4).value  = "Note Number"
          ws_CDNURA.cell(row = 1, column = 5).value  = "Note Date"
          ws_CDNURA.cell(row = 1, column = 6).value  = "Invoice Number"
          ws_CDNURA.cell(row = 1, column = 7).value  = "Invoice Date"
          ws_CDNURA.cell(row = 1, column = 8).value  = "Note Type"
          ws_CDNURA.cell(row = 1, column = 9).value  = "Is Pre GST?"
          ws_CDNURA.cell(row = 1, column = 10).value  = "Note Value"
          ws_CDNURA.cell(row = 1, column = 11).value  = "Taxable value"
          ws_CDNURA.cell(row = 1, column = 12).value  = "Rate"
          ws_CDNURA.cell(row = 1, column = 13).value  = "IGST"
          ws_CDNURA.cell(row = 1, column = 14).value  = "CGST"
          ws_CDNURA.cell(row = 1, column = 15).value  = "SGST"
          ws_CDNURA.cell(row = 1, column = 16).value  = "CESS"
          ws_CDNURA.cell(row = 1, column = 17).value  = "Dealer GSTIN"
          ws_CDNURA.cell(row = 1, column = 18).value  = "Filing Period"
          ws_AT = workbook.create_sheet("AT")
          ws_AT.cell(row = 1, column = 1).value  = "POS"
          ws_AT.cell(row = 1, column = 2).value  = "Type"
          ws_AT.cell(row = 1, column = 3).value  = "Gross Advance"
          ws_AT.cell(row = 1, column = 4).value  = "Rate"
          ws_AT.cell(row = 1, column = 5).value  = "IGST"
          ws_AT.cell(row = 1, column = 6).value  = "CGST"
          ws_AT.cell(row = 1, column = 7).value  = "SGST"
          ws_AT.cell(row = 1, column = 8).value  = "CESS"
          ws_AT.cell(row = 1, column = 9).value  = "Dealer GSTIN"
          ws_AT.cell(row = 1, column = 10).value  = "Filing Period"

          ws_ATA = workbook.create_sheet("ATA")
          ws_ATA.cell(row = 1, column = 1).value  = "Original Month"
          ws_ATA.cell(row = 1, column = 2).value  = "Place of Supply"
          ws_ATA.cell(row = 1, column = 3).value  = "Type"
          ws_ATA.cell(row = 1, column = 4).value  = "Gross Advance"
          ws_ATA.cell(row = 1, column = 5).value  = "Rate"
          ws_ATA.cell(row = 1, column = 6).value  = "IGST"
          ws_ATA.cell(row = 1, column = 7).value  = "CGST"
          ws_ATA.cell(row = 1, column = 8).value  = "SGST"
          ws_ATA.cell(row = 1, column = 9).value  = "CESS"
          ws_ATA.cell(row = 1, column = 10).value  = "Dealer GSTIN"
          ws_ATA.cell(row = 1, column = 11).value  = "Filing Period"

          ws_DOCS = workbook.create_sheet("DOCS")
          ws_DOCS.cell(row = 1, column = 1).value  = "Nature Of Document"
          ws_DOCS.cell(row = 1, column = 2).value  = "Sr No From"
          ws_DOCS.cell(row = 1, column = 3).value  = "Sr No To"
          ws_DOCS.cell(row = 1, column = 4).value  = "Total Number"
          ws_DOCS.cell(row = 1, column = 5).value  = "Canceled"
          ws_DOCS.cell(row = 1, column = 6).value  = "Net"
          ws_DOCS.cell(row = 1, column = 7).value  = "Dealer GSTIN"
          ws_DOCS.cell(row = 1, column = 8).value  = "Filing Period"
          ws_EXEMP = workbook.create_sheet("EXEMP")
          ws_EXEMP.cell(row = 1, column = 1).value  = "Description"
          ws_EXEMP.cell(row = 1, column = 2).value  = "Nil Rated Supplies"
          ws_EXEMP.cell(row = 1, column = 3).value  = "Exempted Supplies"
          ws_EXEMP.cell(row = 1, column = 4).value  = "Non GST Supplies"
          ws_EXEMP.cell(row = 1, column = 5).value  = "Dealer GSTIN"
          ws_EXEMP.cell(row = 1, column = 6).value  = "Filing Period"
          ws_ATADJ = workbook.create_sheet("ATADJ")
          ws_ATADJ.cell(row = 1, column = 1).value  = "Place of Supply"
          ws_ATADJ.cell(row = 1, column = 2).value  = "Supply Type"
          ws_ATADJ.cell(row = 1, column = 3).value  = "Gross Advance Adjusted"
          ws_ATADJ.cell(row = 1, column = 4).value  = "Rate"
          ws_ATADJ.cell(row = 1, column = 5).value  = "IGST"
          ws_ATADJ.cell(row = 1, column = 6).value  = "CGST"
          ws_ATADJ.cell(row = 1, column = 7).value  = "SGST"
          ws_ATADJ.cell(row = 1, column = 8).value  = "CESS"
          ws_ATADJ.cell(row = 1, column = 9).value  = "Dealer GSTIN"
          ws_ATADJ.cell(row = 1, column = 10).value  = "Filing Period"






          ws_ATADJA = workbook.create_sheet("ATADJA")
          ws_ATADJA.cell(row = 1, column = 1).value  = "Place of Supply"
          ws_ATADJA.cell(row = 1, column = 2).value  = "Supply Type"
          ws_ATADJA.cell(row = 1, column = 3).value  = "Gross Advance Adjusted"
          ws_ATADJA.cell(row = 1, column = 4).value  = "Rate"
          ws_ATADJA.cell(row = 1, column = 5).value  = "IGST"
          ws_ATADJA.cell(row = 1, column = 6).value  = "CGST"
          ws_ATADJA.cell(row = 1, column = 7).value  = "SGST"
          ws_ATADJA.cell(row = 1, column = 8).value  = "CESS"
          ws_ATADJA.cell(row = 1, column = 9).value  = "Original Month"
          ws_ATADJA.cell(row = 1, column = 10).value  = "Dealer GSTIN"
          ws_ATADJA.cell(row = 1, column = 11).value  = "Filing Period"


       


          
          
          
          





          ws_HSN = workbook.create_sheet("HSN")
          ws_HSN.cell(row = 1, column = 1).value  = "S No"
          ws_HSN.cell(row = 1, column = 2).value  = "HSN Code"
          ws_HSN.cell(row = 1, column = 3).value  = "Description" 
          ws_HSN.cell(row = 1, column = 4).value  = "UQC"
          ws_HSN.cell(row = 1, column = 5).value  = "QTY"
          ws_HSN.cell(row = 1, column = 6).value  = "Value"
          ws_HSN.cell(row = 1, column = 7).value  = "Taxable Value"
          ws_HSN.cell(row = 1, column = 8).value  = "IGST"
          ws_HSN.cell(row = 1, column = 9).value  = "CGST"
          ws_HSN.cell(row = 1, column = 10).value  = "SGST"
          ws_HSN.cell(row = 1, column = 11).value  = "CESS"
          ws_HSN.cell(row = 1, column = 12).value  = "Dealer GSTIN"
          ws_HSN.cell(row = 1, column = 13).value  = "Filing Period"


          

          row101 = 1
          row102 = 1
          rowb2b = 2
          rowb2cl = 2
          rowb2ba = 2
          rowb2cla = 2
          rowb2cs = 2
          rowb2csa = 2
          rowexp = 2
          rowhsn = 2
          rowexpa = 2 
          rowcdnr = 2
          rowcdnra = 2
          rowcdnur = 2
          rowcdnura = 2
          rowat = 2
          rowata = 2
          rowdocs = 2
          rowexemp = 2
          rowatadj = 2
          rowatadja = 2
          r_count = 0
          
          
          
          
          
          

     
               
               
      # unzip the zip file to the same directory 
          if form.is_valid(): 
               for f in files:
                   # handle_uploaded_file(f)
                    with zipfile.ZipFile(f, 'r') as zip_ref:
                         first = zip_ref.infolist()[0]
                         with zip_ref.open(first, "r") as fo:
                              a = json.load(fo)
          
          #a = json.loads(data)
                         if form.is_valid():
                              try:
                                   for key in a.keys():
                                        if isinstance(a[key], dict)== False:
                                             ws_Info.cell(row = row101, column = 1).value = (key) 
                                             ws_Info.cell(row = row101, column = 2).value = (a['gstin']) 
                                             ws_Info.cell(row = row101, column = 3).value = (a['fp'])
                                             ws_Info.cell(row = row101, column = 4).value = datetime.now()
                                             row101 += 1
                              except:
                                   pass
                              try: 
                                   ws_Info.cell(row = row102, column = 5).value = "B2B"
                              except:
                                   pass
                              try:
                                   ws_Info.cell(row = row102, column = 6).value = (len(a['b2b']))
                              except:
                                   pass
                              try:
                                   ws_Info.cell(row = row102, column = 7).value = (a['gstin']) 
                                   ws_Info.cell(row = row102, column = 8).value = (a['fp'])
                                   ws_Info.cell(row = row102, column = 9).value = (a['fil_dt'])
                                   

                                   row102 += 1
                              except:
                                   pass
                              
                              ws_Info.cell(row = row102, column = 5).value = "B2CL"
                              try:
                                   ws_Info.cell(row = row102, column = 6).value = (len(a['b2cl']))
                              except:
                                   pass
                              try:
                                   ws_Info.cell(row = row102, column = 7).value = (a['gstin']) 
                                   ws_Info.cell(row = row102, column = 8).value = (a['fp'])
                                   row102 += 1
                              except:
                                   pass
                              ws_Info.cell(row = row102, column = 5).value = "B2BA"
                              try:
                                   ws_Info.cell(row = row102, column = 6).value = (len(a['b2ba']))
                              except:
                                   pass
                              ws_Info.cell(row = row102, column = 7).value = (a['gstin']) 
                              ws_Info.cell(row = row102, column = 8).value = (a['fp'])
                              row102 += 1
                              ws_Info.cell(row = row102, column = 5).value = "B2CLA"
                              try:
                                   ws_Info.cell(row = row102, column = 6).value = (len(a['b2cla']))
                              except:
                                   pass
                              ws_Info.cell(row = row102, column = 7).value = (a['gstin']) 
                              ws_Info.cell(row = row102, column = 8).value = (a['fp'])
                              row102 += 1
                              
                              ws_Info.cell(row = row102, column = 5).value = "B2CS"
                              try:
                                   ws_Info.cell(row = row102, column = 6).value = (len(a['b2cs']))
                              except:
                                   pass
                              ws_Info.cell(row = row102, column = 7).value = (a['gstin']) 
                              ws_Info.cell(row = row102, column = 8).value = (a['fp'])
                              row102 += 1

                              ws_Info.cell(row = row102, column = 5).value = "B2CSA"
                              try:
                                   
                                        

                                   ws_Info.cell(row = row102, column = 6).value = (len(a['b2csa'])) , "s"
                              
                                             
                              except:
                                   pass
                              ws_Info.cell(row = row102, column = 7).value = (a['gstin']) 
                              ws_Info.cell(row = row102, column = 8).value = (a['fp'])
                              row102 += 1
                              ws_Info.cell(row = row102, column = 5).value = "EXP1"
                              row102 += 1
                              ws_Info.cell(row = row102, column = 5).value = "EXP2"
                              try:
                                   ws_Info.cell(row = row102, column = 6).value = (len(a['exp'][0]['inv']))
                                   row102 += 1
                                   ws_Info.cell(row = row102, column = 6).value = (len(a['exp'][1]['inv']))

                              except:
                                   pass
                              ws_Info.cell(row = row102, column = 7).value = (a['gstin']) 
                              ws_Info.cell(row = row102, column = 8).value = (a['fp'])
                              row102 += 1

                              ws_Info.cell(row = row102, column = 5).value = "HSN"
                              try:
                                   ws_Info.cell(row = row102, column = 6).value = (len(a['hsn']['data']))
                              except:
                                   pass
                              ws_Info.cell(row = row102, column = 7).value = (a['gstin']) 
                              ws_Info.cell(row = row102, column = 8).value = (a['fp'])    
                              row102 += 1

                              ws_Info.cell(row = row102, column = 5).value = "DOCS"
                              try:
                                   ws_Info.cell(row = row102, column = 6).value = (len(a['doc_issue']['doc_det']))
                              except:
                                   pass
                              ws_Info.cell(row = row102, column = 7).value = (a['gstin']) 
                              ws_Info.cell(row = row102, column = 8).value = (a['fp'])    
                              row102 += 1



 
                                        #ws_Info.cell(row = 4, column = 4).value = (len(a['b2cla']))    
                              #worksheet['A2'] = len(a['b2b'])
                              i = 0 # stands for number of GSTIN in B2B records
                               # Stands for Row 2 indicates data write is gonna start from 2nd row
                              k = 0 # Stands for count of invoices in a GSTIN Record
                              l = 0 # Stands for number of invoice line items in a invoice record
          
                              try:
                                   while i < (len(a['b2b'])):
                                        while k < len(a['b2b'][i]['inv']):
                                             while l < (len(a['b2b'][i]['inv'][k]['itms'])):
                                                  #worksheet.cell(row = rowb2b, column = 10).value = (a['b2b'][i]['inv'][k]['itms'][l]['num'])
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 8).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['rt'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 9).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['txval'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 10).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['iamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 11).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['camt'])
                                                  except:
                                                       pass        
                                                  try: 
                                                       worksheet.cell(row = rowb2b, column = 12).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['samt'])
                                                  except:
                                                       pass 
                                                  try: 
                                                       worksheet.cell(row = rowb2b, column = 13).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['csamt'])
                                                  except:
                                                       pass   
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 1).value = (a['b2b'][i]['ctin'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 2).value = (a['b2b'][i]['inv'][k]['val'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 3).value = (a['b2b'][i]['inv'][k]['inv_typ'])
                                                  except:   
                                                       pass
                                                  try: 
                                                       worksheet.cell(row = rowb2b, column = 4).value = (a['b2b'][i]['inv'][k]['pos'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 5).value = (a['b2b'][i]['inv'][k]['idt'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 6).value = (a['b2b'][i]['inv'][k]['rchrg'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 7).value = (a['b2b'][i]['inv'][k]['inum'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 14).value = (a['gstin'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = rowb2b, column = 15).value = (a['fp'])
                                                  except:
                                                       pass
                                                  
                                                  r_count += 1
                                                  l += 1 # Refers to callout the next invoice level line item hope it starts with 0
                                                  rowb2b += 1 # Excel offset move to next row    
                                             l = 0 # Resetting to 0 for a new record 
                                             k += 1 # Refers to callout next invoice item for a gst record
                                        i += 1 # Moving to next GSTIN
                                        k = 0 # Resetting to 0 for a new record of Invoice
                                   
                              except:
                                   pass

                              try:
                                   sa = 0
                                   
                                   while sa < (len(a['b2cl'])):

                                        sb = 0    
                                        while sb < (len(a['b2cl'][sa]['inv'])):
                                             sc = 0
                                             while sc < (len(a['b2cl'][sa]['inv'][0]['itms'])):
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 5).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['rt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 6).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['txval'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 7).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['iamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 8).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['camt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 9).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['samt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 10).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['csamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 1).value = (a['b2cl'][sa]['inv'][sb]['inum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 2).value = (a['b2cl'][sa]['inv'][sb]['idt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 3).value = (a['b2cl'][sa]['inv'][sb]['val'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 12).value = (a['b2cl'][sa]['inv'][sb]['inv_typ'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 11).value = (a['b2cl'][sa]['inv'][sb]['diff_percent'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 4).value = (a['b2cl'][sa]['pos'])  
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 13).value = (a['gstin'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = rowb2cl, column = 14).value = (a['fp'])
                                                  except:
                                                       pass
                                                  r_count += 1
                                                  rowb2cl += 1
                                                  sc += 1
                                             sb += 1
                                             sc = 0
                                        sb = 0
                                        sc = 0         
                                        sa += 1 
                              except:
                                   pass
                              
                              
                              try:
                                   ta = 0
                                   while ta < (len(a['b2ba'])):
                                        tb = 0
                                        while tb < (len(a['b2ba'][ta]['inv'])):
                                             tc = 0
                                             while tc < (len(a['b2ba'][ta]['inv'][tb]['itms'])):
                                                  try: 
                                                       ws_B2BA.cell(row = rowb2ba, column = 1).value  = ((a['b2ba'][ta]['ctin']))
                                                  except: 
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 2).value  = ((a['b2ba'][ta]['inv'][tb]['oinum']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 3).value  = ((a['b2ba'][ta]['inv'][tb]['oidt']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 4).value  = ((a['b2ba'][ta]['inv'][tb]['inum']))
                                                  except:
                                                       pass
                                                  try :
                                                       ws_B2BA.cell(row = rowb2ba, column = 5).value  = ((a['b2ba'][ta]['inv'][tb]['idt']))
                                                  except:  
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 6).value  = ((a['b2ba'][ta]['inv'][tb]['val']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 7).value  = ((a['b2ba'][ta]['inv'][tb]['pos']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 8).value  = ((a['b2ba'][ta]['inv'][tb]['rchrg']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 9).value  = ((a['b2ba'][ta]['inv'][tb]['diff_percent']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 10).value  = ((a['b2ba'][ta]['inv'][tb]['inv_typ']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 11).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['txval']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 12).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['rt']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 13).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['iamt']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 14).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['camt']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 15).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['samt']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 16).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['csamt']))              
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 17).value = (a['gstin'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = rowb2ba, column = 18).value = (a['fp'])
                                                  except:
                                                       pass
                                                  r_count += 1
                                                  rowb2ba += 1
                                                  tc += 1
                                             tb += 1
                                        ta += 1 
                              except:
                                   pass         

                              try: 
                                   ua = 0
                                   
                                   while ua < (len(a['b2cla'])):                         
                                        ub = 0
                                        while ub < (len(a['b2cla'][ua]['inv'])):
                                             uc = 0   
                                             while uc < (len(a['b2cla'][ua]['inv'][ub]['itms'])):
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 1).value  = (a['b2cla'][ua]['pos'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 9).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['txval'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 10).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['rt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 11).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['iamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 12).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['camt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 13).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['samt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 14).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['csamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 2).value  = (a['b2cla'][ua]['inv'][ub]['oinum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 3).value  = (a['b2cla'][ua]['inv'][ub]['oidt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 4).value  = (a['b2cla'][ua]['inv'][ub]['inum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 5).value  = (a['b2cla'][ua]['inv'][ub]['idt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 6).value  = (a['b2cla'][ua]['inv'][ub]['val'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 7).value  = (a['b2cla'][ua]['inv'][ub]['diff_percent'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = rowb2cla, column = 8).value  = (a['b2cla'][ua]['inv'][ub]['inv_typ'])

                                                  except:
                                                       pass
                                                  ws_B2CLA.cell(row = rowb2cla, column = 9).value  = (a['gstin'])
                                                  ws_B2CLA.cell(row = rowb2cla, column = 10).value  = (a['fp'])
                                                  r_count += 1
                                                  
                                                  
                                                  
                                                  
                                                  rowb2cla += 1
                                                  uc += 1
                                             ub += 1
                                        ua += 1  
                              except:
                                   pass
                              try:
                                   va = 0
                                   
                                   while va < (len(a['b2cs'])):
                                        try:
                                             ws_B2CS.cell(row = rowb2cs, column = 1).value  = (a['b2cs'][va]['sply_ty'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = rowb2cs, column = 2).value  = (a['b2cs'][va]['rt'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = rowb2cs, column = 3).value  = (a['b2cs'][va]['typ'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = rowb2cs, column = 4).value  = (a['b2cs'][va]['pos'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = rowb2cs, column = 5).value  = (a['b2cs'][va]['diff_percent'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = rowb2cs, column = 6).value  = (a['b2cs'][va]['txval'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = rowb2cs, column = 7).value  = (a['b2cs'][va]['iamt'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = rowb2cs, column = 8).value  = (a['b2cs'][va]['camt'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = rowb2cs, column = 9).value  = (a['b2cs'][va]['samt'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = rowb2cs, column = 10).value  = (a['b2cs'][va]['csamt'])
                                        except:
                                             pass
                                        ws_B2CS.cell(row = rowb2cs, column = 11).value  = (a['gstin'])
                                        ws_B2CS.cell(row = rowb2cs, column = 12).value  = (a['fp'])
                                        va += 1
                                        rowb2cs += 1
                                        r_count +=1
                              except:
                                   pass
                              try:
                                   wa = 0
                                   
                                   while wa < (len(a['b2csa'])):
                                        wb = 0
                                        while wb < (len(a['b2csa'][wa]['itms'])):
                                             try:
                                                  ws_B2CSA.cell(row = rowb2csa, column = 1).value  = (a['b2csa'][wa]['omon'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = rowb2csa, column = 2).value  = (a['b2csa'][wa]['sply_ty'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = rowb2csa, column = 3).value  = (a['b2csa'][wa]['typ'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = rowb2csa, column = 4).value  = (a['b2csa'][wa]['pos'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = rowb2csa, column = 5).value  = (a['b2csa'][wa]['diff_percent'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = rowb2csa, column = 6).value  = (a['b2csa'][wa]['itms'][wb]['txval'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = rowb2csa, column = 7).value  = (a['b2csa'][wa]['itms'][wb]['rt'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = rowb2csa, column = 8).value  = (a['b2csa'][wa]['itms'][wb]['iamt'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = rowb2csa, column = 9).value  = (a['b2csa'][wa]['itms'][wb]['camt'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = rowb2csa, column = 10).value  = (a['b2csa'][wa]['itms'][wb]['samt'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = rowb2csa, column = 11).value  = (a['b2csa'][wa]['itms'][wb]['csamt'])
                                             except:
                                                  pass
                                             ws_B2CSA.cell(row = rowb2csa, column = 12).value  = (a['gstin'])
                                             ws_B2CSA.cell(row = rowb2csa, column = 13).value  = (a['fp'])
                                             r_count += 1
                                             rowb2csa += 1
                                             wb += 1
                                        wa += 1
                              except:
                                   pass
                              
                              
                              try:
                                   xa = 0
                                   while xa < (len(a['exp'])):
                                        xb = 0
                                        while xb < (len(a['exp'][xa]['inv'])):
                                             xc = 0
                                             while xc < (len(a['exp'][xa]['inv'][xb]['itms'])):
                                                  try:
                                                       ws_EXP.cell(row = rowexp, column = 1).value  = (a['exp'][xa]['exp_typ'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = rowexp, column = 2).value  = (a['exp'][xa]['inv'][xb]['inum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = rowexp, column = 3).value  = (a['exp'][xa]['inv'][xb]['idt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = rowexp, column = 4).value  = (a['exp'][xa]['inv'][xb]['val'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = rowexp, column = 5).value  = (a['exp'][xa]['inv'][xb]['sbpcode'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = rowexp, column = 6).value  = (a['exp'][xa]['inv'][xb]['sbnum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = rowexp, column = 7).value  = (a['exp'][xa]['inv'][xb]['sbdt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = rowexp, column = 8).value  = (a['exp'][xa]['inv'][xb]['itms'][xc]['txval'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = rowexp, column = 9).value  = (a['exp'][xa]['inv'][xb]['itms'][xc]['rt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = rowexp, column = 10).value  = (a['exp'][xa]['inv'][xb]['itms'][xc]['iamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = rowexp, column = 11).value  = (a['exp'][xa]['inv'][xb]['itms'][xc]['csamt'])
                                                  except:
                                                       pass
                                                  ws_EXP.cell(row = rowexp, column = 12).value  = (a['gstin'])
                                                  ws_EXP.cell(row = rowexp, column = 13).value  = (a['fp'])
                                                  r_count += 1
                                                  rowexp += 1
                                                  xc += 1
                                             xb += 1
                                        xa += 1
                              except:
                                   pass   
                                 
                              
                              try:
                                   za = 0
                                   while za < (len(a['hsn']['data'])): 
                                        try: 
                                             ws_HSN.cell(row = rowhsn, column = 1).value  = (a['hsn']['data'][za]['num'])
                                        except:
                                             pass
                                        try:     
                                             ws_HSN.cell(row = rowhsn, column = 2).value  = (a['hsn']['data'][za]['hsn_sc'])
                                        except:
                                             pass
                                        try:     
                                             ws_HSN.cell(row = rowhsn, column = 3).value  = (a['hsn']['data'][za]['desc'])
                                        except:
                                             pass
                                        try:     
                                             ws_HSN.cell(row = rowhsn, column = 4).value  = (a['hsn']['data'][za]['uqc'])
                                        except:
                                             pass
                                        try:          
                                             ws_HSN.cell(row = rowhsn, column = 5).value  = (a['hsn']['data'][za]['qty'])
                                        except:
                                             pass
                                        try:          
                                             ws_HSN.cell(row = rowhsn, column = 6).value  = (a['hsn']['data'][za]['val'])
                                        except:
                                             pass
                                        try:          
                                             ws_HSN.cell(row = rowhsn, column = 7).value  = (a['hsn']['data'][za]['txval'])
                                        except:
                                             pass
                                        try:                                   
                                             ws_HSN.cell(row = rowhsn, column = 8).value  = (a['hsn']['data'][za]['iamt'])
                                        except:
                                             pass
                                        try:                                   
                                             ws_HSN.cell(row = rowhsn, column = 9).value  = (a['hsn']['data'][za]['camt'])
                                        except:
                                             pass
                                        try:          
                                             ws_HSN.cell(row = rowhsn, column = 10).value  = (a['hsn']['data'][za]['samt'])
                                        except:
                                             pass
                                        try:          
                                             ws_HSN.cell(row = rowhsn, column = 11).value  = (a['hsn']['data'][za]['csamt'])
                                        except:
                                             pass
                                        ws_HSN.cell(row = rowhsn, column = 12).value  = (a['gstin'])
                                        ws_HSN.cell(row = rowhsn, column = 13).value  = (a['fp'])
                                        rowhsn += 1
                                        r_count +=1
                                        za += 1  

                              except:
                                   pass  
                              
                              try:
                                   ya = 0
                                   while ya < (len(a['expa'])):
                                        yb = 0
                                        while yb < (len(a['expa'][ya]['inv'])):
                                             yc = 0
                                             while yc < (len(a['expa'][ya]['inv'][yb]['itms'])):
                                                  try:
                                                       ws_EXPA.cell(row = rowexpa, column = 1).value  = (a['expa'][ya]['exp_typ'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = rowexpa, column = 2).value  = (a['expa'][ya]['inv'][yb]['inum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = rowexpa, column = 3).value  = (a['expa'][ya]['inv'][yb]['idt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = rowexpa, column = 4).value  = (a['expa'][ya]['inv'][yb]['val'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = rowexpa, column = 5).value  = (a['expa'][ya]['inv'][yb]['sbpcode'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = rowexpa, column = 6).value  = (a['expa'][ya]['inv'][yb]['sbnum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = rowexpa, column = 7).value  = (a['expa'][ya]['inv'][yb]['sbdt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = rowexpa, column = 8).value  = (a['expa'][ya]['inv'][yb]['itms'][yc]['txval'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = rowexpa, column = 9).value  = (a['expa'][ya]['inv'][yb]['itms'][yc]['rt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = rowexpa, column = 10).value  = (a['expa'][ya]['inv'][yb]['itms'][yc]['iamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = rowexpa, column = 11).value  = (a['expa'][ya]['inv'][yb]['itms'][yc]['csamt'])
                                                  except:
                                                       pass
                                                  ws_EXPA.cell(row = rowexpa, column = 12).value  = (a['expa'][ya]['inv'][yb]['oinum'])
                                                  ws_EXPA.cell(row = rowexpa, column = 13).value  = (a['expa'][ya]['inv'][yb]['oidt'])
                                                  ws_EXPA.cell(row = rowexpa, column = 14).value =(a['gstin'])
                                                  ws_EXPA.cell(row = rowexpa, column = 15).value = (a['fp'])
                                                  
                                                  r_count += 1
                                                  rwoexpa += 1
                                                  yc += 1
                                             yb += 1
                                        ya += 1
                              except:
                                   pass                         

                                   try:
                                        
                                        aa = 0
                                        
                                        
                                        while aa < (len(a['cdnr'])):
                                             ab = 0
                                             while ab < (len(a['cdnr'][aa]['nt'])):
                                                  ac = 0
                                                  while ac < (len(a['cdnr'][aa]['nt'][ab]['itms'])):
                                                  #     try:
                                                       
                                                       ws_CDNR.cell(row = rowcdnr, column = 1).value  = (a['cdnr'][aa]['ctin'])
                                                       ws_CDNR.cell(row = rowcdnr, column = 2).value  = (a['cdnr'][aa]['nt'][ab]['val'])
                                                       ws_CDNR.cell(row = rowcdnr, column = 3).value  = (a['cdnr'][aa]['nt'][ab]['ntty'])
                                                       ws_CDNR.cell(row = rowcdnr, column = 4).value  = (a['cdnr'][aa]['nt'][ab]['nt_num'])
                                                       ws_CDNR.cell(row = rowcdnr, column = 5).value  = (a['cdnr'][aa]['nt'][ab]['inum'])
                                                       ws_CDNR.cell(row = rowcdnr, column = 6).value  = (a['cdnr'][aa]['nt'][ab]['idt'])
                                                       
                                                       ws_CDNR.cell(row = rowcdnr, column = 7).value  = (a['cdnr'][aa]['nt'][ab]['nt_dt'])
                                                       ws_CDNR.cell(row = rowcdnr, column = 8).value  = (a['cdnr'][aa]['nt'][ab]['p_gst'])
                                                       
                                                       ws_CDNR.cell(row = rowcdnr, column = 9).value  = (a['cdnr'][aa]['nt'][ab]['itms'][ac]['itm_det']['txval'])
                                                       try:
                                                            ws_CDNR.cell(row = rowcdnr, column = 10).value  = (a['cdnr'][aa]['nt'][ab]['itms'][ac]['itm_det']['rt'])
                                                       except:
                                                            pass
                                                       try:
                                                            ws_CDNR.cell(row = rowcdnr, column = 11).value  = (a['cdnr'][aa]['nt'][ab]['itms'][ac]['itm_det']['iamt'])   
                                                       except:
                                                            pass
                                                       try:
                                                            ws_CDNR.cell(row = rowcdnr, column = 12).value  = (a['cdnr'][aa]['nt'][ab]['itms'][ac]['itm_det']['camt'])
                                                       except:
                                                            pass
                                                       
                                                            
                                                      
                                                      
                                                       try:
                                                            ws_CDNR.cell(row = rowcdnr, column = 13).value  = (a['cdnr'][aa]['nt'][ab]['itms'][ac]['itm_det']['samt'])
                                                       except:
                                                            pass
                                                       try:
                                                            ws_CDNR.cell(row = rowcdnr, column = 14).value  = (a['cdnr'][aa]['nt'][ab]['itms'][ac]['itm_det']['csamt'])
                                                            ws_CDNR.cell(row = rowcdnr, column = 15).value  = (a['gstin'])
                                                            ws_CDNR.cell(row = rowcdnr, column = 16).value  = (a['fp'])
                                                       except:
                                                            pass
                                                       rowcdnr += 1
                                                  
                                                  
                                                       ac += 1
                                                  ab += 1
                                             aa += 1
                                                  
                                             

                                   except:
                                        pass
                                        
                                 
                                   try:
     
                                        ba = 0


                                        while ba < (len(a['cdnra'])): # cdnra count
                                             bb = 0
                                             while bb < (len(a['cdnra'][ba]['nt'])): # no  of Notes count
                                                  bc = 0
                                                  while ac < (len(a['cdnra'][ba]['nt'][bb]['itms'])):
                                                       try:
                                                            ws_CDNRA.cell(row = rowcdnra, column = 1).value  = (a['cdnra'][ba]['ctin'])
                                                       except:
                                                            pass
                                                       try: 
                                                            ws_CDNRA.cell(row = rowcdnra, column = 2).value  = (a['cdnra'][ba]['nt'][bb]['val'])
                                                       except:
                                                            pass
                                                       try:
                                                            ws_CDNRA.cell(row = rowcdnra, column = 3).value  = (a['cdnra'][ba]['nt'][bb]['ntty'])
                                                       except:
                                                            pass
                                                       try:
                                                            ws_CDNRA.cell(row = rowcdnra, column = 4).value  = (a['cdnra'][ba]['nt'][bb]['ont_num'])
                                                       except:
                                                            pass
                                                       try: 
                                                            ws_CDNRA.cell(row = rowcdnra, column = 5).value  = (a['cdnra'][ba]['nt'][bb]['ont_dt'])
                                                       except:
                                                            pass
                                                       try:
                                                            ws_CDNRA.cell(row = rowcdnra, column = 6).value  = (a['cdnra'][ba]['nt'][bb]['nt_num'])
                                                       except:
                                                            pass
                                                       try: 
                                                            ws_CDNRA.cell(row = rowcdnra, column = 7).value  = (a['cdnra'][ba]['nt'][bb]['inum'])
                                                       except:
                                                            pass
                                                       try:
                                                            ws_CDNRA.cell(row = rowcdnra, column = 8).value  = (a['cdnra'][ba]['nt'][bb]['idt'])
                                                       except:
                                                            pass
                                                       ws_CDNRA.cell(row = rowcdnra, column = 9).value  = (a['cdnra'][ba]['nt'][bb]['nt_dt'])
                                                       ws_CDNRA.cell(row = rowcdnra, column = 10).value  = (a['cdnra'][ba]['nt'][bb]['p_gst'])

                                                       ws_CDNRA.cell(row = rowcdnra, column = 11).value  = (a['cdnra'][ba]['nt'][bb]['itms'][bc]['itm_det']['txval'])
                                                       try:
                                                            ws_CDNRA.cell(row = rowcdnra, column = 12).value  = (a['cdnra'][ba]['nt'][bb]['itms'][bc]['itm_det']['rt'])
                                                       except:
                                                            pass
                                                       try:
                                                            ws_CDNRA.cell(row = rowcdnra, column = 13).value  = (a['cdnra'][ba]['nt'][bb]['itms'][bc]['itm_det']['iamt'])   
                                                       except:
                                                            pass
                                                       try:
                                                            ws_CDNRA.cell(row = rowcdnra, column = 14).value  = (a['cdnra'][ba]['nt'][bb]['itms'][bc]['itm_det']['camt'])
                                                       except:
                                                            pass


                                                       try:
                                                            ws_CDNRA.cell(row = rowcdnra, column = 15).value  = (a['cdnra'][ba]['nt'][bb]['itms'][bc]['itm_det']['samt'])
                                                       except:
                                                            pass
                                                       try:
                                                            ws_CDNRA.cell(row = rowcdnra, column = 16).value  = (a['cdnra'][ba]['nt'][bb]['itms'][bc]['itm_det']['csamt'])
                                                       except:
                                                            pass
                                                       
                                                       try:     
                                                            ws_CDNRA.cell(row = rowcdnra, column = 17).value  = (a['gstin'])
                                                       except:
                                                            pass
                                                       try:
                                                            
                                                            ws_CDNRA.cell(row = rowcdnra, column = 18).value  = (a['fp'])
                                                       except:
                                                            pass
                                                       rowcdnra += 1
                                                       bc += 1
                                                  bb +=1
                                             ba += 1
                                   except:
                                        pass        
                                   try:
                                        ca = 0
                                        while ca < (len(a['cdnur'])): 
                                             cb = 0
                                             while cb < (len(a['cdnur'][ca]['itms'])):
                                             
                                             
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 1).value  = (a['cdnur'][ca]['typ'])         
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 2).value  = (a['cdnur'][ca]['nt_num'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 3).value  = (a['cdnur'][ca]['nt_dt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 4).value  = (a['cdnur'][ca]['inum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 5).value  = (a['cdnur'][ca]['idt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 6).value  = (a['cdnur'][ca]['ntty'])
                                                  except:
                                                       pass
                                                  try: 
                                                       ws_CDNUR.cell(row = rowcdnur, column = 7).value  = (a['cdnur'][ca]['p_gst'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 8).value  = (a['cdnur'][ca]['val'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 9).value  = (a['cdnur'][ca]['itms'][cb]['itm_det']['txval'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 10).value  = (a['cdnur'][ca]['itms'][cb]['itm_det']['rt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 11).value  = (a['cdnur'][ca]['itms'][cb]['itm_det']['iamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 12).value  = (a['cdnur'][ca]['itms'][cb]['itm_det']['camt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 13).value  = (a['cdnur'][ca]['itms'][cb]['itm_det']['samt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 14).value  = (a['cdnur'][ca]['itms'][cb]['itm_det']['csamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 15).value  = (a['gstin'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNUR.cell(row = rowcdnur, column = 16).value  = (a['fp'])
                                                  except:
                                                       pass
                                                       #ws_CDNUR.cell(row = 2, column = 1).value  = (a['cdnur'][ca]['nt_num'])


                                                  r_count += 1
                                                  rowcdnur += 1 
                                                  
                                                  cb += 1

                                             ca += 1
                                   except:
                                        pass                                                        
                                                                                                    
                                   try:
                                        da = 0
                                        while da < (len(a['cdnura'])): 
                                             db = 0
                                             while db < (len(a['cdnura'][da]['itms'])):
                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 1).value  = (a['cdnura'][da]['typ'])  
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 2).value  = (a['cdnura'][da]['ont_num'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 3).value  = (a['cdnura'][da]['ont_dt'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 4).value  = (a['cdnura'][da]['nt_num'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 5).value  = (a['cdnura'][da]['nt_dt'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 6).value  = (a['cdnura'][da]['inum'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 7).value  = (a['cdnura'][da]['idt'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 8).value  = (a['cdnura'][da]['ntty'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 9).value  = (a['cdnura'][da]['p_gst'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 10).value  = (a['cdnura'][da]['val'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 11).value  = (a['cdnura'][da]['itms'][db]['itm_det']['txval'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 12).value  = (a['cdnura'][da]['itms'][db]['itm_det']['rt'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 13).value  = (a['cdnura'][da]['itms'][db]['itm_det']['iamt'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 14).value  = (a['cdnura'][da]['itms'][db]['itm_det']['camt'])
                                                  except:
                                                       pass

                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 15).value  = (a['cdnura'][da]['itms'][db]['itm_det']['samt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_CDNURA.cell(row = rowcdnura, column = 16).value  = (a['cdnura'][da]['itms'][db]['itm_det']['csamt'])
                                                  except:
                                                       pass
                                                  ws_CDNURA.cell(row = rowcdnura, column = 17).value  = (a['gstin'])
                                                  ws_CDNURA.cell(row = rowcdnura, column = 18).value  = (a['fp'])
                                                  
                                                  r_count += 1
                                                  rowcdnura += 1
                                                  db += 1
                                             da += 1

                                   except:
                                        pass       
                                   try:
                                        ea = 0
                                        while ea < (len(a['at'])):
                                             eb = 0
                                             while eb < (len(a['at'][ea]['itms'])):

                                                                          
                                                  try:
                                                       ws_AT.cell(row = rowat, column =1).value  = (a['at'][ea]['pos'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_AT.cell(row = rowat, column =2).value  = (a['at'][ea]['sply_ty'])    
                                                  except:
                                                       pass
                                                  try:
                                                       ws_AT.cell(row = rowat, column =3).value  = (a['at'][ea]['itms'][eb]['ad_amt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_AT.cell(row = rowat, column =4).value  = (a['at'][ea]['itms'][eb]['rt'])    
                                                  except:
                                                       pass
                                                  try:
                                                       ws_AT.cell(row = rowat, column =5).value  = (a['at'][ea]['itms'][eb]['iamt'])  
                                                  except:
                                                       pass
                                                  try:
                                                       ws_AT.cell(row = rowat, column =6).value  = (a['at'][ea]['itms'][eb]['camt'])
                                                  except:
                                                       pass
                                                  try:     
                                                       ws_AT.cell(row = rowat, column =7).value  = (a['at'][ea]['itms'][eb]['samt'])  
                                                  except:
                                                       pass     
                                                  try:     
                                                       ws_AT.cell(row = rowat, column =8).value  = (a['at'][ea]['itms'][eb]['csamt'])    
                                                  except:
                                                       pass 
                                                  try:     
                                                       ws_AT.cell(row = rowat, column =9).value  = (a['gstin']) 
                                                  except:
                                                       pass 
                                                  try:     
                                                       ws_AT.cell(row = rowat, column =10).value  = (a['fp']) 
                                                  except:
                                                       pass 


                                                  rowat += 1
                                                  r_count += 1
                                                  eb += 1
                                             ea += 1

                                   except:
                                        pass     
                                   try:
                                        fa = 0
                                        while fa < (len(a['ata'])):
                                             fb = 0
                                             while fb < (len(a['ata'][fa]['itms'])):
                                                  #try:
                                                  ws_ATA.cell(row = rowata, column =1).value  = (a['ata'][fa]['omon']) 
                                                  #except:
                                                  #     pass
                                                  try:
                                                       ws_ATA.cell(row = rowata, column =2).value  = (a['ata'][fa]['pos'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_ATA.cell(row = rowata, column =3).value  = (a['ata'][fa]['sply_ty'])    
                                                  except:
                                                       pass
                                                  try:
                                                       ws_ATA.cell(row = rowata, column =4).value  = (a['ata'][fa]['itms'][fb]['ad_amt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_ATA.cell(row = rowata, column =5).value  = (a['ata'][fa]['itms'][fb]['rt'])    
                                                  except:
                                                       pass
                                                  try:
                                                       ws_ATA.cell(row = rowata, column =6).value  = (a['ata'][fa]['itms'][fb]['iamt'])  
                                                  except:
                                                       pass
                                                  try:
                                                       ws_ATA.cell(row = rowata, column =7).value  = (a['ata'][fa]['itms'][fb]['camt'])
                                                  except:
                                                       pass
                                                  try:     
                                                       ws_ATA.cell(row = rowata, column =8).value  = (a['ata'][fa]['itms'][fb]['samt'])  
                                                  except:
                                                       pass     
                                                  try:     
                                                       ws_ATA.cell(row = rowata, column =9).value  = (a['ata'][fa]['itms'][fb]['csamt'])    
                                                  except:
                                                       pass 
                                                  try:     
                                                       ws_ATA.cell(row = rowata, column =10).value  = (a['gstin']) 
                                                  except:
                                                       pass 
                                                  try:     
                                                       ws_ATA.cell(row = rowata, column =11).value  = (a['fp']) 
                                                  except:
                                                       pass 
                                                  rowata += 1
                                                  r_count += 1
                                                  fb += 1
                                             fa += 1
                                   except:
                                        pass                 

                                   try:
                                        ga = 0
                                        
                                        while ga < (len(a['doc_issue']['doc_det'])):
                                             gb = 0
                                             while gb < (len(a['doc_issue']['doc_det'][ga]['docs'])):
                                                  
                                                  
                                                  
                                                  
                                                  try:
                                                       ws_DOCS.cell(row = rowdocs, column =1).value  = (a['doc_issue']['doc_det'][ga]['doc_typ'])
                                                  except: 
                                                       pass
                                                  
                                                  
                                                  
                                                  
                                                  try:   
                                                       ws_DOCS.cell(row = rowdocs, column =2).value  = (a['doc_issue']['doc_det'][ga]['docs'][gb]['from'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_DOCS.cell(row = rowdocs, column =3).value  = (a['doc_issue']['doc_det'][ga]['docs'][gb]['to'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_DOCS.cell(row = rowdocs, column =4).value  = (a['doc_issue']['doc_det'][ga]['docs'][gb]['totnum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_DOCS.cell(row = rowdocs, column =5).value  = (a['doc_issue']['doc_det'][ga]['docs'][gb]['cancel'])
                                                  except:
                                                       pass
                                                  try:

                                                       ws_DOCS.cell(row = rowdocs, column =6).value  = (a['doc_issue']['doc_det'][ga]['docs'][gb]['net_issue'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_DOCS.cell(row = rowdocs, column =7).value  = (a['gstin'])
                                                       ws_DOCS.cell(row = rowdocs, column =8).value  = (a['fp'])
                                                  except:
                                                       pass
                                                  rowdocs += 1
                                                  r_count += 1
                                                  gb += 1
                                             ga += 1                  
                                   except:
                                       pass
                                                                
                                                                
                                   try:
                                        ha = 0

                                        while ha < (len(a['nil']['inv'])):
                                             ws_EXEMP.cell(row = rowexemp, column =1).value  = (a['nil']['inv'][ha]['sply_ty']) 
                                             ws_EXEMP.cell(row = rowexemp, column =2).value  = (a['nil']['inv'][ha]['expt_amt'])
                                             ws_EXEMP.cell(row = rowexemp, column =3).value  = (a['nil']['inv'][ha]['nil_amt'])
                                             ws_EXEMP.cell(row = rowexemp, column =4).value  = (a['nil']['inv'][ha]['ngsup_amt'])
                                             ws_EXEMP.cell(row = rowexemp, column =5).value  = (a['gstin'])
                                             ws_EXEMP.cell(row = rowexemp, column =6).value  = (a['fp'])
                                             r_count += 1
                                             rowexemp += 1
                                             ha +=1
                                   except:
                                        pass                             
                                                                
                                   try:
                                        ia = 0
                                        while ia < (len(a['txpd'])):
                                             ib = 0
                                             while ib < (len(a['txpd'][ga]['itms'])):
                                                  ws_ATADJ.cell(row = rowatadj, column =1).value  = (a['txpd'][ia]['pos'])
                                                  ws_ATADJ.cell(row = rowatadj, column =2).value  = (a['txpd'][ia]['sply_ty'])
                                                  ws_ATADJ.cell(row = rowatadj, column =3).value  = (a['txpd'][ia]['itms'][ib]['ad_amt'])
                                                  ws_ATADJ.cell(row = rowatadj, column =4).value  = (a['txpd'][ia]['itms'][ib]['rt'])
                                                  try:
                                                       ws_ATADJ.cell(row = rowatadj, column =5).value  = (a['txpd'][ia]['itms'][ib]['iamt'])
                                                  except:
                                                       pass
                                                  try:     
                                                       ws_ATADJ.cell(row = rowatadj, column =6).value  = (a['txpd'][ia]['itms'][ib]['camt'])
                                                  except:
                                                       pass
                                                  try:     
                                                       ws_ATADJ.cell(row = rowatadj, column =7).value  = (a['txpd'][ia]['itms'][ib]['samt'])
                                                  except:
                                                       pass
                                                  try:     
                                                       ws_ATADJ.cell(row = rowatadj, column =8).value  = (a['txpd'][ia]['itms'][ib]['csamt'])
                                                  except:
                                                       pass
                                                  ws_ATADJ.cell(row = rowatadj, column =9).value  = (a['gstin'])
                                                  ws_ATADJ.cell(row = rowatadj, column =10).value  = (a['fp'])
                                                  rowatadj += 1
                                                  ib  += 1
                                                  r_count += 1
                                             ia +=1
                                   except:
                                        pass                             
                             
                             
                              try:
                                   ja = 0
                                   while ja < (len(a['txpda'])):
                                        jb = 0
                                        while jb < (len(a['txpda'][ja]['itms'])):
                                             ws_ATADJA.cell(row = rowatadja, column =1).value  = (a['txpda'][ja]['pos'])
                                             ws_ATADJA.cell(row = rowatadja, column =2).value  = (a['txpda'][ja]['sply_ty'])
                                             ws_ATADJA.cell(row = rowatadja, column =3).value  = (a['txpda'][ja]['itms'][jb]['ad_amt'])
                                             ws_ATADJA.cell(row = rowatadja, column =4).value  = (a['txpda'][ja]['itms'][jb]['rt'])
                                             try:
                                                  ws_ATADJA.cell(row = rowatadja, column =5).value  = (a['txpda'][ja]['itms'][jb]['iamt'])
                                             except:
                                                  pass
                                             try:     
                                                  ws_ATADJA.cell(row = rowatadja, column =6).value  = (a['txpda'][ja]['itms'][jb]['camt'])
                                             except:
                                                  pass
                                             try:     
                                                  ws_ATADJA.cell(row = rowatadja, column =7).value  = (a['txpda'][ja]['itms'][jb]['samt'])
                                             except:
                                                  pass
                                             try:     
                                                  ws_ATADJA.cell(row = rowatadja, column =8).value  = (a['txpda'][ja]['itms'][jb]['csamt'])
                                                  
                                             except:
                                                  pass
                                             ws_ATADJA.cell(row = rowatadja, column =9).value  = (a['txpda'][ja]['omon'])
                                             ws_ATADJA.cell(row = rowatadja, column =10).value  = (a['gstin'])
                                             ws_ATADJA.cell(row = rowatadja, column =11).value  = (a['fp'])
                                             rowatadja += 1
                                             jb  += 1
                                             r_count += 1
                                        ja +=1
                              except:
                                   pass                             
                                  
                                  
                                  
                                  
                                  
                                  
                                  
                                  
                                  



                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                             #                             
                                                                
                                                                
                                                                
                                                                
                                                                
                                                                
                                                                
                                                                
                                                                
                                                                
                                                                
                                                                

                                                                
                                                                


                              




                              



               myobject = Gstworker(GSTIN=(a['gstin']), r_counts=(r_count))
               myobject.save()               
               workbook.save(response)
          return response
     else:  
          student = StudentForm()  
          return render(request,"index.html",{'form':student})              
                               
                              













                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
                              
               
                              
                              


















































































































































































































































































































































































































































































































































































































































































